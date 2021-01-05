# Главная основная форма приложения
# Выбор файла БД и режима дальнейшей работы
import sqlite3
import sys
from urllib.parse import quote

import xlsxwriter
from PyQt5 import uic
from PyQt5.QtCore import Qt
from PyQt5.QtSql import QSqlDatabase, QSqlRelationalTableModel, QSqlRelation, \
    QSqlRelationalDelegate
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QTableWidgetItem, QFileDialog, \
    QMessageBox, QDataWidgetMapper
from openpyxl import load_workbook

from db_view import Ui_DB_View_Form
from editForm import Ui_EditForm
from importForm import Ui_ImportForm


class Database:
    def __init__(self, db_filename):
        self.connection = sqlite3.connect(db_filename)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # БД по умолчанию
        self.DB_Name = 'inventarizaciya10.db'
        uic.loadUi('mainWindow.ui', self)
        self.importWinBtn.clicked.connect(self.run_import_from)
        self.ViewdbBtn.clicked.connect(self.run_db_view)
        self.print_int_btn.clicked.connect(self.run_print_inv_form)
        self.file_open_btn.clicked.connect(self.openfile)

    def openfile(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Выбор базы данных", "", "DataBase Files (*.db);;All Files (*)")
        if file_name:
            self.db_filename.setText(file_name)
            self.DB_Name = file_name

    def run_db_view(self):
        self.db_view_win = DBViewWindow(self.DB_Name)
        self.db_view_win.show()

    def run_import_from(self):
        self.importForm = ImportForm(self.DB_Name)
        self.importForm.show()

    def run_print_inv_form(self):
        self.printForm = PrintInvForm(self.DB_Name)
        self.printForm.show()


class PrintInvForm(QWidget):
    def __init__(self, database_name):
        super().__init__()
        uic.loadUi('printInvForm.ui', self)
        self.db_name = database_name
        self.save_btn.clicked.connect(self.make_document)
        self.file_open_btn.clicked.connect(self.choose_file)

    def connect_db(self):
        pass

    def choose_file(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Исходные данные", "", "Excel Files (*.xlsx;*xls);;All Files (*)")
        if file_name:
            self.filename.setText(file_name)

    def make_document(self):
        def add_to_file(income_res, worksheet_name, target_worksheet):
            target_worksheet.name = worksheet_name
            for row, data in enumerate(income_res):
                target_worksheet.write(row, 0, data[0])
                target_worksheet.write(row, 1, data[1])
                target_worksheet.write(row, 2, quote(data[1]), code128_format)
                target_worksheet.write(row, 3, data[2])

        connection = sqlite3.connect(self.db_name)
        query = f"SELECT goods_name, invent_number, location_name " \
                f"FROM goods, location " \
                f"WHERE location_id=id_location"
        res = connection.cursor().execute(query).fetchall()
        connection.close()
        dest_filename = self.filename.text()
        if dest_filename:
            workbook = xlsxwriter.Workbook(dest_filename)
            code128_format = workbook.add_format({'font_name': 'Code 128'})
            ws1 = workbook.add_worksheet()
            add_to_file(res, "Все позиции", ws1)
            # далее делаем разбивку по location на отдельные листы
            dif_locations = set([data[2] for data in res])
            for location in dif_locations:
                connection = sqlite3.connect(self.db_name)
                query = f"SELECT goods_name, invent_number, location_name " \
                        f"FROM goods, location " \
                        f"WHERE location_id=id_location AND location_name='{location}'"
                res = connection.cursor().execute(query).fetchall()
                connection.close()
                add_to_file(res, location, workbook.add_worksheet())
            workbook.close()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Успешно!")
            msg.setText(f"Сохранение в \n{dest_filename}\nпрошло успешно")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("ОШИБКА!")
            msg.setText(f"Нет имени файла!!!")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()


class EditForm(QWidget, Ui_EditForm):
    def __init__(self, database, current_index, model):
        super().__init__()
        # uic.loadUi('editForm.ui', self)
        self.setupUi(self)
        self.db = database

        self.db_map = QDataWidgetMapper(self)
        self.db_map.setModel(model)
        self.db_map.setSubmitPolicy(QDataWidgetMapper.ManualSubmit)
        self.db_map.addMapping(self.lineEdit_name, 1)
        self.db_map.addMapping(self.lineEdit_inv_number, 2)
        self.db_map.addMapping(self.textEdit_comments, 3)
        self.db_map.addMapping(self.checkBox_on_balance, model.fieldIndex('is_on_balance'))

        relModel_status = model.relationModel(5)
        self.comboBox_status.setModel(relModel_status)
        self.comboBox_status.setModelColumn(relModel_status.fieldIndex("status_name"))
        self.db_map.setItemDelegate(QSqlRelationalDelegate(self))
        self.db_map.addMapping(self.comboBox_status, 5)

        relModel_type = model.relationModel(6)
        self.comboBox_type.setModel(relModel_type)
        self.comboBox_type.setModelColumn(relModel_type.fieldIndex("goods_type_name"))
        self.db_map.setItemDelegate(QSqlRelationalDelegate(self))
        self.db_map.addMapping(self.comboBox_type, 6)

        relModel_subtype = model.relationModel(7)
        self.comboBox_subtype.setModel(relModel_subtype)
        self.comboBox_subtype.setModelColumn(relModel_subtype.fieldIndex("goods_subtype_name"))
        self.db_map.setItemDelegate(QSqlRelationalDelegate(self))
        self.db_map.addMapping(self.comboBox_subtype, 7)

        relModel_location = model.relationModel(8)
        self.comboBox_location.setModel(relModel_location)
        self.comboBox_location.setModelColumn(relModel_location.fieldIndex("location_name"))
        self.db_map.setItemDelegate(QSqlRelationalDelegate(self))
        self.db_map.addMapping(self.comboBox_location, 8)

        relModel_responsible = model.relationModel(9)
        self.comboBox_responsible.setModel(relModel_responsible)
        self.comboBox_responsible.setModelColumn(relModel_responsible.fieldIndex("FIO"))
        self.db_map.setItemDelegate(QSqlRelationalDelegate(self))
        self.db_map.addMapping(self.comboBox_responsible, 9)

        self.pushBtn_save.clicked.connect(self.save_item)
        self.pushBtn_cancel.clicked.connect(self.cancel)
        self.pushBtn_next.clicked.connect(self.db_map.toNext)
        self.pushBtn_prev.clicked.connect(self.db_map.toPrevious)
        self.pushBtn_close.clicked.connect(self.close)

        self.db_map.setCurrentIndex(current_index.row())

    def save_item(self):
        self.db_map.submit()

    def cancel(self):
        self.db_map.setCurrentIndex(self.db_map.currentIndex())


class DBViewWindow(QWidget, Ui_DB_View_Form):
    def __init__(self, database_name):
        super().__init__()
        # uic.loadUi('db_view.ui', self)
        self.setupUi(self)
        self.db_name = database_name
        # Подключение БД к таблице отображения
        # Подключение через QSqlRelationalTableModel
        self.db = QSqlDatabase.addDatabase('QSQLITE')
        self.db.setDatabaseName(self.db_name)
        self.db.open()
        self.model = QSqlRelationalTableModel(self)
        self.model.setTable('goods')
        self.model.setRelation(5, QSqlRelation('statuses', 'id_status', 'status_name'))
        self.model.setRelation(6, QSqlRelation('goods_types', 'id_goods_type', 'goods_type_name'))
        self.model.setRelation(7, QSqlRelation('goods_subtypes', 'id_goods_subtype',
                                               'goods_subtype_name'))
        self.model.setRelation(8, QSqlRelation('location', 'id_location', 'location_name'))
        self.model.setRelation(9, QSqlRelation('responsibles', 'id_responsible', 'FIO'))
        self.refresh()
        self.tableView.doubleClicked.connect(self.table_clicked)
        # self.tableView.clicked.connect()
        self.refreshBtn.clicked.connect(self.refresh)
        self.save_all_btn.clicked.connect(self.submitall)

    def table_clicked(self):
        row = self.tableView.currentIndex().row()
        if row != -1:
            self.tableView.selectRow(row)
        index = self.tableView.currentIndex()
        self.edit_form = EditForm(self.db, index, self.model)
        self.edit_form.show()

    # Предположительно на выходе будет закрываться БД
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.db.close()

    def refresh(self):
        self.model.select()
        self.tableView.setModel(self.model)
        self.tableView.setItemDelegate(QSqlRelationalDelegate(self.tableView))

    def submitall(self):
        if not self.model.submitAll():
            print(self.model.lastError())
        else:
            self.refresh()


class ImportForm(QWidget, Ui_ImportForm):
    def __init__(self, database_name):
        super().__init__()
        # uic.loadUi('importForm.ui', self)
        self.setupUi(self)
        self.db_name_edit.setText(database_name)
        self.not_braked = True
        self.db_select_file_btn.clicked.connect(self.get_db_filename)
        self.source_file_btn.clicked.connect(self.get_import_filename)
        self.start_import_button.clicked.connect(self.parse_source)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            msg = QMessageBox()
            msg.setWindowTitle("Запрос")
            msg.setText("Прервать операцию импорта?")
            msg.setInformativeText("")
            msg.setDetailedText("")
            result = msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            retval = msg.exec_()
            if retval == QMessageBox.Yes:
                self.not_braked = False
            else:
                self.not_braked = True

    def get_db_filename(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "БД для куда импортируем")
        if file_name:
            self.db_name_edit.setText(file_name)

    def get_import_filename(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Исходные данные", "", "Excel Files (*.xlsx;*xls);;All Files (*)")
        if file_name:
            self.sheets_list_box.addItem('Loading...')
            self.source_file_edit.setText(file_name)
            self.repaint()
            self.workbook = load_workbook(filename=self.source_file_edit.text())
            self.sheets_list_box.setEnabled(True)
            self.sheets_list_box.clear()
            self.sheets_list_box.addItems(self.workbook.sheetnames)
            self.sheets_list_box.setCurrentIndex(1)  # второй лист по умолчанию

    def parse_source(self):

        the_sheet = self.workbook[self.sheets_list_box.currentText()]
        values = []
        columns = "AGH"
        row = 8
        rows_on_list = 20
        rows_between_lists = 10
        on_list = 0
        self.items_table.setColumnCount(len(columns))
        con = sqlite3.connect(self.db_name_edit.text())
        cursor = con.cursor()
        all_goods = []
        que = ''
        while row <= the_sheet.max_row and self.not_braked:
            if on_list >= rows_on_list:
                on_list = 0
                row += rows_between_lists
            no = the_sheet[f"{columns[0]}{row}"]
            name = the_sheet[f"{columns[1]}{row}"]
            inv_num = the_sheet[f"{columns[2]}{row}"]
            if not inv_num.value:
                # если нет инвентарного номера то
                inv_num.value = "NO_INV_" + str(no.value)
            if no.value and int(no.value) == len(values) + 1:
                values.append((no.value, name.value, inv_num.value))
                self.items_table.setRowCount(self.items_table.rowCount() + 1)
                for i, item in enumerate(values[-1]):
                    self.items_table.setItem(self.items_table.rowCount() - 1, i, QTableWidgetItem(
                        str(item)))
                all_goods.append(tuple([name.value, inv_num.value]))
            else:
                print(f"{no.value} != {len(values) + 1}\n"
                      f"the data is not added:\n"
                      f"{name.value},{inv_num.value}")
            excluded = "excluded"  # только чтобы не подсвечивалась как ошибка
            que = f"INSERT into goods(goods_name, invent_number) " \
                  f"VALUES {', '.join([str(x) for x in all_goods])} " \
                  f"ON CONFLICT (invent_number) DO " \
                  f"UPDATE SET " \
                  f"comment = comment ||'\n Старое значение Наименования'||goods_name, " \
                  f"goods_name = {excluded}.goods_name"
            row += 1
            on_list += 1
            # self.keyPressEvent()
        cursor.execute(que)
        con.commit()
        self.items_table.repaint()
        con.close()
        # print(the_sheet['A28'].value)
        print("Импортировано записей:", len(values))


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWin = MainWindow()
    mainWin.show()
    sys.excepthook = except_hook
    sys.exit(app.exec_())
